from fastapi import APIRouter, Depends, HTTPException, status, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, extract
from typing import List, Optional
from datetime import datetime, date, time, timezone, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.database.connection import get_db
from app.database.models import TimeEntry, Task, User, Project
from app.schemas.time_entry import TimeEntryCreate, TimeEntryResponse, TimeEntryUpdate
from app.core.deps import get_current_user

router = APIRouter(prefix="/time-entries", tags=["Time Entries"])

# ============================================
# UTILITY FUNCTIONS
# ============================================

def get_utc_now():
    """Return datetime NAIVE (no timezone) for DB compatibility"""
    return datetime.utcnow()

def serialize_entry_with_utc(entry, include_task_info=False, db=None):
    """
    Convert TimeEntry to dict with explicit UTC timezone
    ✅ NEW: Optionally include task/project info for frontend
    """
    result = {
        "id": entry.id,
        "task_id": entry.task_id,
        "user_id": entry.user_id,
        "start_time": entry.start_time.isoformat() + "Z" if not entry.start_time.tzinfo else entry.start_time.isoformat(),
        "end_time": (entry.end_time.isoformat() + "Z" if not entry.end_time.tzinfo else entry.end_time.isoformat()) if entry.end_time else None,
        "duration": entry.duration or 0,
        "note": entry.note or ""
    }
    
    # ✅ NEW: Include task/project info if requested
    if include_task_info and db:
        task = db.query(Task).filter(Task.id == entry.task_id).first()
        if task:
            result["task_title"] = task.title
            result["task_status"] = task.status
            
            project = db.query(Project).filter(Project.id == task.project_id).first()
            if project:
                result["project_id"] = project.id
                result["project_name"] = project.name
                result["project_color"] = getattr(project, 'color', None)
    
    return result

def get_date_range_for_period(period: str):
    """
    Get start and end dates for common periods
    ✅ NEW: Support for 'today', 'yesterday', 'this_week', 'last_week', 'this_month', 'last_month'
    """
    today = date.today()
    
    if period == "today":
        return today, today
    
    elif period == "yesterday":
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    
    elif period == "this_week":
        # Monday of current week
        start = today - timedelta(days=today.weekday())
        return start, today
    
    elif period == "last_week":
        # Monday to Sunday of last week
        last_monday = today - timedelta(days=today.weekday() + 7)
        last_sunday = last_monday + timedelta(days=6)
        return last_monday, last_sunday
    
    elif period == "this_month":
        start = today.replace(day=1)
        return start, today
    
    elif period == "last_month":
        first_this_month = today.replace(day=1)
        last_day_last_month = first_this_month - timedelta(days=1)
        first_last_month = last_day_last_month.replace(day=1)
        return first_last_month, last_day_last_month
    
    else:
        raise ValueError(f"Invalid period: {period}")

# ============================================
# 1. START TIMER (ENHANCED)
# ============================================

@router.post("/start")
def start_timer(
    entry_data: TimeEntryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Start a new timer (auto-stops any running timer)
    ✅ ENHANCED: Returns full task/project info
    """
    task = db.query(Task).filter(Task.id == entry_data.task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task không tồn tại")
    
    # Verify user has access to this task
    project = db.query(Project).filter(Project.id == task.project_id).first()
    if not project or project.owner_id != current_user.id:
        raise HTTPException(status_code=403, detail="Không có quyền truy cập task này")

    # Auto-stop running timer
    running_entry = db.query(TimeEntry).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.end_time.is_(None)
    ).first()

    if running_entry:
        now = get_utc_now()
        running_entry.end_time = now
        duration = max(0, int((now - running_entry.start_time).total_seconds()))
        running_entry.duration = duration
        
        old_task = db.query(Task).filter(Task.id == running_entry.task_id).first()
        if old_task:
            old_task.total_time = (old_task.total_time or 0) + duration

    new_entry = TimeEntry(
        task_id=entry_data.task_id,
        user_id=current_user.id,
        note=entry_data.note or "",
        start_time=get_utc_now(),
        end_time=None,
        duration=0
    )
    db.add(new_entry)
    db.commit()
    db.refresh(new_entry)
    
    return serialize_entry_with_utc(new_entry, include_task_info=True, db=db)

# ============================================
# 2. STOP TIMER (ENHANCED)
# ============================================

@router.post("/stop")
def stop_timer(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Stop the currently running timer
    ✅ ENHANCED: Returns full entry info with calculations
    """
    running_entry = db.query(TimeEntry).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.end_time.is_(None)
    ).first()

    if not running_entry:
        raise HTTPException(status_code=400, detail="Không có timer nào đang chạy")

    now = get_utc_now()
    running_entry.end_time = now
    duration = max(0, int((now - running_entry.start_time).total_seconds()))
    running_entry.duration = duration

    task = db.query(Task).filter(Task.id == running_entry.task_id).first()
    if task:
        task.total_time = (task.total_time or 0) + duration

    db.commit()
    db.refresh(running_entry)
    
    result = serialize_entry_with_utc(running_entry, include_task_info=True, db=db)
    result["hours"] = round(duration / 3600, 2)
    
    return result

# ============================================
# 3. GET CURRENT TIMER (UNCHANGED BUT ENHANCED)
# ============================================

@router.get("/current")
def get_current_timer(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get currently running timer
    ✅ ENHANCED: Returns task/project info
    """
    entry = db.query(TimeEntry).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.end_time.is_(None)
    ).first()
    
    if not entry:
        return None
    
    return serialize_entry_with_utc(entry, include_task_info=True, db=db)

# ============================================
# 4. GET HISTORY WITH FILTERS (ENHANCED)
# ============================================

@router.get("/")
def get_my_time_entries(
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    task_id: Optional[int] = None,
    project_id: Optional[int] = None,  # ✅ NEW
    period: Optional[str] = None,  # ✅ NEW: 'today', 'this_week', etc.
    include_running: bool = True,  # ✅ NEW
    include_task_info: bool = False,  # ✅ NEW
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get time entries with comprehensive filtering
    ✅ ENHANCED: More filter options, optimized queries
    """
    query = db.query(TimeEntry).filter(TimeEntry.user_id == current_user.id)
    
    # Handle period shortcuts
    if period:
        try:
            period_start, period_end = get_date_range_for_period(period)
            start_date = period_start
            end_date = period_end
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    # Date filtering
    if start_date:
        dt_start = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time >= dt_start)
    
    if end_date:
        dt_end = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time <= dt_end)
    
    # Task filtering
    if task_id:
        query = query.filter(TimeEntry.task_id == task_id)
    
    # ✅ NEW: Project filtering (requires join)
    if project_id:
        query = query.join(Task).filter(Task.project_id == project_id)
    
    # ✅ NEW: Exclude running entries if requested
    if not include_running:
        query = query.filter(TimeEntry.end_time.isnot(None))

    total = query.count()
    skip = (page - 1) * limit
    
    entries = query.order_by(TimeEntry.start_time.desc())\
                   .offset(skip)\
                   .limit(limit)\
                   .all()
    
    # Serialize with optional task info
    entries_data = [
        serialize_entry_with_utc(entry, include_task_info=include_task_info, db=db) 
        for entry in entries
    ]
    
    return {
        "data": entries_data,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit,
        "has_next": skip + limit < total
    }

# ============================================
# 5. GET SUMMARY (ENHANCED)
# ============================================

@router.get("/summary")
def get_time_summary(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    period: Optional[str] = None,  # ✅ NEW
    group_by: str = Query("task", regex="^(task|project|date)$"),  # ✅ NEW
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get summary statistics with flexible grouping
    ✅ ENHANCED: Support grouping by task, project, or date
    """
    # Handle period shortcuts
    if period:
        try:
            start_date, end_date = get_date_range_for_period(period)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    query = db.query(TimeEntry).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.end_time.isnot(None)
    )
    
    if start_date:
        dt_start = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time >= dt_start)
    
    if end_date:
        dt_end = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time <= dt_end)
    
    entries = query.all()
    
    # Calculate totals
    total_seconds = sum(entry.duration or 0 for entry in entries)
    total_hours = round(total_seconds / 3600, 2)
    
    # Group based on group_by parameter
    if group_by == "task":
        grouped_data = _group_by_task(entries, db)
    elif group_by == "project":
        grouped_data = _group_by_project(entries, db)
    elif group_by == "date":
        grouped_data = _group_by_date(entries)
    else:
        grouped_data = []
    
    return {
        "total_hours": total_hours,
        "total_entries": len(entries),
        "total_seconds": total_seconds,
        "date_range": {
            "start": start_date.isoformat() if start_date else None,
            "end": end_date.isoformat() if end_date else None
        },
        "group_by": group_by,
        "grouped_data": grouped_data
    }

def _group_by_task(entries, db):
    """Group entries by task"""
    task_breakdown = {}
    for entry in entries:
        task_id = entry.task_id
        if task_id not in task_breakdown:
            task = db.query(Task).filter(Task.id == task_id).first()
            project = db.query(Project).filter(Project.id == task.project_id).first() if task else None
            
            task_breakdown[task_id] = {
                "task_id": task_id,
                "task_name": task.title if task else f"Task #{task_id}",
                "project_id": project.id if project else None,
                "project_name": project.name if project else None,
                "total_seconds": 0,
                "entry_count": 0
            }
        
        task_breakdown[task_id]["total_seconds"] += entry.duration or 0
        task_breakdown[task_id]["entry_count"] += 1
    
    tasks = []
    for task_data in task_breakdown.values():
        tasks.append({
            **task_data,
            "total_hours": round(task_data["total_seconds"] / 3600, 2)
        })
    
    tasks.sort(key=lambda x: x["total_hours"], reverse=True)
    return tasks

def _group_by_project(entries, db):
    """Group entries by project"""
    project_breakdown = {}
    
    for entry in entries:
        task = db.query(Task).filter(Task.id == entry.task_id).first()
        if not task:
            continue
        
        project_id = task.project_id
        if project_id not in project_breakdown:
            project = db.query(Project).filter(Project.id == project_id).first()
            project_breakdown[project_id] = {
                "project_id": project_id,
                "project_name": project.name if project else f"Project #{project_id}",
                "project_color": project.color if project else None,
                "total_seconds": 0,
                "entry_count": 0,
                "task_count": set()
            }
        
        project_breakdown[project_id]["total_seconds"] += entry.duration or 0
        project_breakdown[project_id]["entry_count"] += 1
        project_breakdown[project_id]["task_count"].add(entry.task_id)
    
    projects = []
    for project_data in project_breakdown.values():
        projects.append({
            "project_id": project_data["project_id"],
            "project_name": project_data["project_name"],
            "project_color": project_data["project_color"],
            "total_hours": round(project_data["total_seconds"] / 3600, 2),
            "entry_count": project_data["entry_count"],
            "task_count": len(project_data["task_count"])
        })
    
    projects.sort(key=lambda x: x["total_hours"], reverse=True)
    return projects

def _group_by_date(entries):
    """Group entries by date"""
    date_breakdown = {}
    
    for entry in entries:
        entry_date = entry.start_time.date().isoformat()
        
        if entry_date not in date_breakdown:
            date_breakdown[entry_date] = {
                "date": entry_date,
                "total_seconds": 0,
                "entry_count": 0
            }
        
        date_breakdown[entry_date]["total_seconds"] += entry.duration or 0
        date_breakdown[entry_date]["entry_count"] += 1
    
    dates = []
    for date_data in date_breakdown.values():
        dates.append({
            **date_data,
            "total_hours": round(date_data["total_seconds"] / 3600, 2)
        })
    
    dates.sort(key=lambda x: x["date"], reverse=True)
    return dates

# ============================================
# 6. TODAY STATS (NEW - OPTIMIZED)
# ============================================

@router.get("/stats/today")
def get_today_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get quick stats for today
    ✅ NEW: Optimized for dashboard
    """
    today = date.today()
    dt_start = datetime.combine(today, time.min).replace(tzinfo=timezone.utc)
    dt_end = datetime.combine(today, time.max).replace(tzinfo=timezone.utc)
    
    entries = db.query(TimeEntry).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.start_time >= dt_start,
        TimeEntry.start_time <= dt_end
    ).all()
    
    completed_entries = [e for e in entries if e.end_time is not None]
    running_entry = next((e for e in entries if e.end_time is None), None)
    
    total_seconds = sum(e.duration or 0 for e in completed_entries)
    
    # Calculate running timer duration
    running_seconds = 0
    if running_entry:
        running_seconds = int((get_utc_now() - running_entry.start_time).total_seconds())
    
    return {
        "date": today.isoformat(),
        "total_hours": round((total_seconds + running_seconds) / 3600, 2),
        "completed_hours": round(total_seconds / 3600, 2),
        "total_entries": len(entries),
        "completed_entries": len(completed_entries),
        "has_running_timer": running_entry is not None,
        "running_task_id": running_entry.task_id if running_entry else None,
        "running_seconds": running_seconds
    }

# ============================================
# 7. CALENDAR DATA (NEW)
# ============================================

@router.get("/calendar")
def get_calendar_data(
    year: int = Query(..., ge=2020, le=2100),
    month: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Get time entry data for calendar view
    ✅ NEW: Support calendar integration
    """
    # Get first and last day of month
    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)
    
    dt_start = datetime.combine(first_day, time.min).replace(tzinfo=timezone.utc)
    dt_end = datetime.combine(last_day, time.max).replace(tzinfo=timezone.utc)
    
    entries = db.query(TimeEntry).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.start_time >= dt_start,
        TimeEntry.start_time <= dt_end,
        TimeEntry.end_time.isnot(None)
    ).all()
    
    # Group by date
    calendar_data = {}
    for entry in entries:
        entry_date = entry.start_time.date().isoformat()
        
        if entry_date not in calendar_data:
            calendar_data[entry_date] = {
                "date": entry_date,
                "total_seconds": 0,
                "total_hours": 0,
                "entry_count": 0,
                "tasks": set()
            }
        
        calendar_data[entry_date]["total_seconds"] += entry.duration or 0
        calendar_data[entry_date]["entry_count"] += 1
        calendar_data[entry_date]["tasks"].add(entry.task_id)
    
    # Convert to list and calculate hours
    result = []
    for day_data in calendar_data.values():
        result.append({
            "date": day_data["date"],
            "total_hours": round(day_data["total_seconds"] / 3600, 2),
            "entry_count": day_data["entry_count"],
            "task_count": len(day_data["tasks"])
        })
    
    return {
        "year": year,
        "month": month,
        "days": result
    }

# ============================================
# 8. BILLING CALCULATIONS (NEW)
# ============================================

@router.get("/billing")
def calculate_billing(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    period: Optional[str] = None,
    project_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Calculate billing for time entries
    ✅ NEW: Support billing calculations with hourly rates
    """
    # Handle period shortcuts
    if period:
        try:
            start_date, end_date = get_date_range_for_period(period)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    # Build query
    query = db.query(TimeEntry, Task, Project).join(
        Task, TimeEntry.task_id == Task.id
    ).join(
        Project, Task.project_id == Project.id
    ).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.end_time.isnot(None)
    )
    
    if start_date:
        dt_start = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time >= dt_start)
    
    if end_date:
        dt_end = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time <= dt_end)
    
    if project_id:
        query = query.filter(Project.id == project_id)
    
    entries = query.all()
    
    # Calculate billing by project
    project_billing = {}
    total_hours = 0
    total_amount = 0
    
    for entry, task, project in entries:
        hours = (entry.duration or 0) / 3600
        rate = project.hourly_rate or 0
        amount = hours * rate
        
        if project.id not in project_billing:
            project_billing[project.id] = {
                "project_id": project.id,
                "project_name": project.name,
                "hourly_rate": rate,
                "total_hours": 0,
                "total_amount": 0,
                "entry_count": 0
            }
        
        project_billing[project.id]["total_hours"] += hours
        project_billing[project.id]["total_amount"] += amount
        project_billing[project.id]["entry_count"] += 1
        
        total_hours += hours
        total_amount += amount
    
    # Convert to list
    projects = []
    for billing_data in project_billing.values():
        projects.append({
            "project_id": billing_data["project_id"],
            "project_name": billing_data["project_name"],
            "hourly_rate": billing_data["hourly_rate"],
            "total_hours": round(billing_data["total_hours"], 2),
            "total_amount": round(billing_data["total_amount"], 2),
            "entry_count": billing_data["entry_count"]
        })
    
    projects.sort(key=lambda x: x["total_amount"], reverse=True)
    
    return {
        "period": {
            "start": start_date.isoformat() if start_date else None,
            "end": end_date.isoformat() if end_date else None
        },
        "total_hours": round(total_hours, 2),
        "total_amount": round(total_amount, 2),
        "projects": projects
    }

# ============================================
# 9. UPDATE ENTRY (UNCHANGED)
# ============================================

@router.put("/{entry_id}")
def update_time_entry(
    entry_id: int,
    update_data: TimeEntryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update time entry (mainly for editing notes)"""
    entry = db.query(TimeEntry).filter(
        TimeEntry.id == entry_id, 
        TimeEntry.user_id == current_user.id
    ).first()
    
    if not entry:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")
    
    if update_data.note is not None:
        entry.note = update_data.note
    
    db.commit()
    db.refresh(entry)
    return serialize_entry_with_utc(entry)

# ============================================
# 10. DELETE ENTRY (UNCHANGED)
# ============================================

@router.delete("/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_time_entry(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete a time entry"""
    entry = db.query(TimeEntry).filter(
        TimeEntry.id == entry_id, 
        TimeEntry.user_id == current_user.id
    ).first()
    
    if not entry:
        raise HTTPException(status_code=404, detail="Không tìm thấy bản ghi")
    
    # Update task total_time
    if entry.duration and entry.duration > 0:
        task = db.query(Task).filter(Task.id == entry.task_id).first()
        if task and task.total_time:
            task.total_time = max(0, task.total_time - entry.duration)
    
    db.delete(entry)
    db.commit()
    return None

# ============================================
# 11. BULK DELETE (UNCHANGED)
# ============================================

@router.post("/bulk-delete", status_code=status.HTTP_200_OK)
def bulk_delete_entries(
    entry_ids: List[int],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Delete multiple time entries at once"""
    if not entry_ids:
        raise HTTPException(status_code=400, detail="Danh sách rỗng")
    
    entries = db.query(TimeEntry).filter(
        TimeEntry.id.in_(entry_ids),
        TimeEntry.user_id == current_user.id
    ).all()
    
    if not entries:
        raise HTTPException(status_code=404, detail="Không tìm thấy entries")
    
    deleted_count = 0
    for entry in entries:
        if entry.duration and entry.duration > 0:
            task = db.query(Task).filter(Task.id == entry.task_id).first()
            if task and task.total_time:
                task.total_time = max(0, task.total_time - entry.duration)
        
        db.delete(entry)
        deleted_count += 1
    
    db.commit()
    return {
        "message": f"Đã xóa {deleted_count} entries",
        "deleted_count": deleted_count
    }

# ============================================
# 12. EXPORT CSV (ENHANCED)
# ============================================

@router.get("/export/csv")
def export_csv(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    period: Optional[str] = None,  # ✅ NEW
    project_id: Optional[int] = None,  # ✅ NEW
    include_billing: bool = False,  # ✅ NEW
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export time entries to CSV file
    ✅ ENHANCED: Support billing, period shortcuts
    """
    # Handle period shortcuts
    if period:
        try:
            start_date, end_date = get_date_range_for_period(period)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    query = db.query(TimeEntry, Task, Project).join(
        Task, TimeEntry.task_id == Task.id
    ).join(
        Project, Task.project_id == Project.id
    ).filter(
        TimeEntry.user_id == current_user.id
    )
    
    if start_date:
        dt_start = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time >= dt_start)
    
    if end_date:
        dt_end = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time <= dt_end)
    
    if project_id:
        query = query.filter(Project.id == project_id)
    
    entries = query.order_by(TimeEntry.start_time.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # BOM for UTF-8
    output.write('\ufeff')
    
    # Headers
    if include_billing:
        writer.writerow(['Date', 'Start Time', 'End Time', 'Duration (hours)', 'Project', 'Task', 'Note', 'Rate', 'Amount'])
    else:
        writer.writerow(['Date', 'Start Time', 'End Time', 'Duration (hours)', 'Project', 'Task', 'Note'])
    
    total_hours = 0
    total_amount = 0
    
    for entry, task, project in entries:
        if not entry.end_time:
            continue
        
        start = entry.start_time.strftime('%Y-%m-%d')
        start_time = entry.start_time.strftime('%H:%M:%S')
        end_time = entry.end_time.strftime('%H:%M:%S')
        duration_hours = round(entry.duration / 3600, 2) if entry.duration else 0
        
        total_hours += duration_hours
        
        row = [
            start,
            start_time,
            end_time,
            duration_hours,
            project.name,
            task.title,
            entry.note or ''
        ]
        
        if include_billing:
            rate = project.hourly_rate or 0
            amount = duration_hours * rate
            total_amount += amount
            row.extend([rate, round(amount, 2)])
        
        writer.writerow(row)
    
    # Summary row
    writer.writerow([])
    if include_billing:
        writer.writerow(['TOTAL', '', '', total_hours, '', '', '', '', round(total_amount, 2)])
    else:
        writer.writerow(['TOTAL', '', '', total_hours, '', '', ''])
    
    output.seek(0)
    
    filename = f"timesheet_{current_user.username}_{datetime.now().strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============================================
# 13. EXPORT EXCEL (ENHANCED)
# ============================================

@router.get("/export/excel")
def export_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    period: Optional[str] = None,  # ✅ NEW
    project_id: Optional[int] = None,  # ✅ NEW
    include_billing: bool = False,  # ✅ NEW
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Export time entries to Excel file
    ✅ ENHANCED: Better formatting, billing support
    """
    # Handle period shortcuts
    if period:
        try:
            start_date, end_date = get_date_range_for_period(period)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    
    query = db.query(TimeEntry, Task, Project).join(
        Task, TimeEntry.task_id == Task.id
    ).join(
        Project, Task.project_id == Project.id
    ).filter(
        TimeEntry.user_id == current_user.id
    )
    
    if start_date:
        dt_start = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time >= dt_start)
    
    if end_date:
        dt_end = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time <= dt_end)
    
    if project_id:
        query = query.filter(Project.id == project_id)
    
    entries = query.order_by(TimeEntry.start_time.desc()).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    if include_billing:
        headers = ['Date', 'Start', 'End', 'Hours', 'Project', 'Task', 'Note', 'Rate', 'Amount']
    else:
        headers = ['Date', 'Start', 'End', 'Hours', 'Project', 'Task', 'Note']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    total_hours = 0
    total_amount = 0
    
    for row_idx, (entry, task, project) in enumerate(entries, 2):
        if not entry.end_time:
            continue
        
        duration_hours = round(entry.duration / 3600, 2) if entry.duration else 0
        total_hours += duration_hours
        
        ws.cell(row_idx, 1, entry.start_time.strftime('%Y-%m-%d')).border = border
        ws.cell(row_idx, 2, entry.start_time.strftime('%H:%M')).border = border
        ws.cell(row_idx, 3, entry.end_time.strftime('%H:%M')).border = border
        ws.cell(row_idx, 4, duration_hours).border = border
        ws.cell(row_idx, 5, project.name).border = border
        ws.cell(row_idx, 6, task.title).border = border
        ws.cell(row_idx, 7, entry.note or '').border = border
        
        if include_billing:
            rate = project.hourly_rate or 0
            amount = duration_hours * rate
            total_amount += amount
            
            ws.cell(row_idx, 8, rate).border = border
            ws.cell(row_idx, 9, round(amount, 2)).border = border
    
    # Summary row
    summary_row = len(entries) + 2
    summary_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    
    ws.cell(summary_row, 1, "TOTAL").font = summary_font
    ws.cell(summary_row, 1).fill = summary_fill
    ws.cell(summary_row, 4, round(total_hours, 2)).font = summary_font
    ws.cell(summary_row, 4).fill = summary_fill
    
    if include_billing:
        ws.cell(summary_row, 9, round(total_amount, 2)).font = summary_font
        ws.cell(summary_row, 9).fill = summary_fill
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"timesheet_{current_user.username}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ============================================
# 14. STATISTICS BY TASK (UNCHANGED)
# ============================================

@router.get("/stats/by-task")
def get_stats_by_task(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get statistics grouped by task"""
    query = db.query(
        Task.id,
        Task.title,
        Project.name.label('project_name'),
        func.sum(TimeEntry.duration).label('total_seconds'),
        func.count(TimeEntry.id).label('entry_count')
    ).join(
        Task, TimeEntry.task_id == Task.id
    ).join(
        Project, Task.project_id == Project.id
    ).filter(
        TimeEntry.user_id == current_user.id,
        TimeEntry.end_time.isnot(None)
    )
    
    if start_date:
        dt_start = datetime.combine(start_date, time.min).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time >= dt_start)
    
    if end_date:
        dt_end = datetime.combine(end_date, time.max).replace(tzinfo=timezone.utc)
        query = query.filter(TimeEntry.start_time <= dt_end)
    
    results = query.group_by(Task.id, Task.title, Project.name)\
                   .order_by(func.sum(TimeEntry.duration).desc())\
                   .all()
    
    total_seconds = sum(r.total_seconds or 0 for r in results)
    
    stats = []
    for r in results:
        seconds = r.total_seconds or 0
        hours = round(seconds / 3600, 2)
        percentage = round((seconds / total_seconds * 100), 1) if total_seconds > 0 else 0
        
        stats.append({
            "task_id": r.id,
            "task_title": r.title,
            "project_name": r.project_name,
            "total_hours": hours,
            "entry_count": r.entry_count,
            "percentage": percentage
        })
    
    return {
        "stats": stats,
        "total_hours": round(total_seconds / 3600, 2),
        "total_entries": sum(r.entry_count for r in results)
    }